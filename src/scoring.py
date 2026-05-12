"""Pure functions for CSV/xlsx parsing, URL parsing, and report generation.

No side effects — all I/O is done via arguments and return values.
"""

import csv
import logging
import re
from pathlib import Path
from urllib.parse import unquote

import openpyxl

from models import (
    CodeQualityResult,
    ConceptScoreResult,
    DifficultyScoreEntry,
    GroupInfo,
)

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# URL / project-link parsing
# ---------------------------------------------------------------------------


def parse_project_link(link: str) -> tuple[str | None, str | None, bool, bool]:
    """Parse a GitHub project URL into (branch, path, is_zip, is_commit).

    Handles patterns:
      - tree/<branch>/<path>        → (branch, path, False, False)
      - tree/<commit_hash>/<path>   → (commit_hash, path, False, True)
      - blob/<branch>/<path>.zip    → (branch, path, True, False)
      - commit/<hash>               → (hash, None, False, True)
      - empty / None                → (None, None, False, False)

    Args:
        link: Raw URL from the scorecard CSV.

    Returns:
        Tuple of (branch, path, is_zip, is_commit).
    """
    if not link or not link.strip():
        return (None, None, False, False)

    link = link.strip()

    # commit/<hash>
    m = re.search(r"/commit/([0-9a-f]{7,40})", link)
    if m:
        return (m.group(1), None, False, True)

    # blob/<branch>/<path>.zip
    m = re.search(r"/blob/([^/]+)/(.+\.zip)$", link)
    if m:
        branch = unquote(m.group(1))
        path = unquote(m.group(2))
        return (branch, path, True, False)

    # tree/<ref>/<path>  or  tree/<ref>
    m = re.search(r"/tree/([^/]+)(?:/(.+))?$", link)
    if m:
        ref = unquote(m.group(1))
        path = unquote(m.group(2)) if m.group(2) else None

        # Check if ref looks like a commit hash (40 hex chars)
        is_commit = bool(re.fullmatch(r"[0-9a-f]{7,40}", ref))
        return (ref, path, False, is_commit)

    # blob/<branch>/<path>  (non-zip blob, e.g. a folder link mis-categorised)
    m = re.search(r"/blob/([^/]+)/(.+)$", link)
    if m:
        branch = unquote(m.group(1))
        path = unquote(m.group(2))
        return (branch, path, False, False)

    return (None, None, False, False)


def _extract_github_repo_url(link: str) -> str | None:
    """Normalise any GitHub link to a bare clone URL.

    Handles full HTTPS URLs (with or without .git suffix and sub-paths) and
    bare ``user/repo`` partial paths.

    Args:
        link: Raw project link from the scorecard.

    Returns:
        Normalised ``https://github.com/user/repo`` URL, or None if the link
        cannot be mapped to a GitHub repository.
    """
    if not link or not link.strip():
        return None
    link = link.strip()

    # Partial "user/repo" format — no scheme, exactly one slash
    if not link.startswith("http") and link.count("/") == 1:
        return f"https://github.com/{link}"

    # Full GitHub URL — strip everything after the repo name
    m = re.match(r"(https?://github\.com/[^/]+/[^/]+?)(?:\.git)?(?:/.*)?$", link)
    if m:
        return m.group(1)

    return None


def _parse_group_number(raw: object) -> int | None:
    """Extract an integer group number from a raw cell value.

    Handles floats (``1.0`` → ``1``) and strings like ``'17 (Individual)'``.

    Args:
        raw: Raw cell value from the xlsx.

    Returns:
        Integer group number, or None if unparseable.
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == "Group":
        return None
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1))
    return None


# ---------------------------------------------------------------------------
# CSV loading
# ---------------------------------------------------------------------------


def load_groups_from_csv(csv_path: Path) -> list[GroupInfo]:
    """Parse the scorecard CSV into a list of GroupInfo objects.

    Args:
        csv_path: Path to the CSV file (C3 or C4).

    Returns:
        List of GroupInfo, one per group row that has a numeric group number.
    """
    groups: list[GroupInfo] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_group = row.get("Group", "").strip()
            if not raw_group or not raw_group.isdigit():
                continue

            project_link = row.get("Project Link", "").strip()
            video_link = row.get("Video Link", "").strip()
            branch, path, is_zip, is_commit = parse_project_link(project_link)

            groups.append(
                GroupInfo(
                    group=int(raw_group),
                    project_link=project_link,
                    video_link=video_link,
                    branch=branch,
                    path=path,
                    is_zip=is_zip,
                    is_commit=is_commit,
                )
            )

    return groups


def load_groups_from_xlsx(xlsx_path: Path, sheet_name: str) -> list[GroupInfo]:
    """Parse a scorecard xlsx sheet into a list of GroupInfo objects.

    Used for cohorts (e.g. C6) where each group submitted an external GitHub
    repository rather than a branch of a shared local repo.

    Args:
        xlsx_path: Path to the xlsx workbook.
        sheet_name: Name of the sheet to read (e.g. 'C6').

    Returns:
        List of GroupInfo, one per data row with a parseable group number.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    groups: list[GroupInfo] = []

    for row in ws.iter_rows(values_only=True):
        gn = _parse_group_number(row[0])
        if gn is None:
            continue

        project_link = str(row[1]).strip() if row[1] is not None else ""
        # Video Link column may contain ArrayFormula objects when data_only=True
        raw_video = row[2]
        video_link = str(raw_video).strip() if isinstance(raw_video, str) else ""

        # Use standard branch/path extraction for links that have them
        branch, path, is_zip, is_commit = parse_project_link(project_link)

        repo_url = _extract_github_repo_url(project_link)

        # For a bare repo URL parse_project_link returns branch=None; treat as main
        if branch is None and repo_url:
            branch = "main"

        groups.append(
            GroupInfo(
                group=gn,
                project_link=project_link,
                video_link=video_link,
                branch=branch,
                path=path,
                is_zip=is_zip,
                is_commit=is_commit,
                is_external=True,
                external_repo_url=repo_url,
            )
        )

    wb.close()
    return groups


def load_syllabus(csv_path: Path) -> str:
    """Load the syllabus CSV and format it as a readable text block.

    Args:
        csv_path: Path to the syllabus CSV.

    Returns:
        Formatted string summarising each sprint's topics and outcomes.
    """
    lines: list[str] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sprint_title = row.get("Sprint Title", "").strip()
            topics = row.get("Topics", "").strip()
            description = row.get("Description", "").strip()
            outcomes = row.get("Outcomes", "").strip()
            tools = row.get("Tools - Sprint Wise", "").strip()

            if not sprint_title and not topics:
                continue

            block = []
            if sprint_title:
                block.append(f"## {sprint_title}")
            if topics:
                block.append(f"**Topics:** {topics}")
            if description:
                block.append(f"**Description:** {description}")
            if outcomes:
                block.append(f"**Outcomes:** {outcomes}")
            if tools:
                block.append(f"**Tools:** {tools}")
            block.append("")

            lines.append("\n".join(block))

    return "\n".join(lines)


def load_c3_reference(csv_path: Path) -> str:
    """Load C3 scores and format as a reference table for calibration.

    Args:
        csv_path: Path to the C3 scorecard CSV.

    Returns:
        Formatted reference table string.
    """
    lines: list[str] = [
        "# C3 Reference Scores (for calibration)",
        "",
        "| Group | Concept | Difficulty | Code Quality | Total | Comments |",
        "|-------|---------|------------|--------------|-------|----------|",
    ]

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            grp = row.get("Group", "").strip()
            if not grp or not grp.isdigit():
                continue
            concept = row.get("Concept Score (10)", "").strip() or "-"
            diff = row.get("Difficulty Level (10)", "").strip() or "-"
            quality = row.get("Code Quality (10)", "").strip() or "-"
            total = row.get("Total (30)", "").strip() or "-"
            comments = row.get("Comments", "").strip() or ""
            lines.append(
                f"| {grp} | {concept} | {diff} | {quality} | {total} | {comments} |"
            )

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Prompt building for per-project evaluation
# ---------------------------------------------------------------------------


def build_project_prompt(group: GroupInfo, repo: str = "c4") -> str:
    """Build the evaluation prompt for a single group, with tool-use hints.

    Uses ``match`` on the group's link characteristics to provide the right
    instructions for Cursor CLI shell and file inspection.

    Args:
        group: Parsed group metadata.
        repo: Which repo this group belongs to ('c3', 'c4', or 'c6').

    Returns:
        Prompt string to send to the judge agents.
    """
    header = f"# Evaluate Group {group.group}\n\n"

    # External repos (C6 style): clone directly from GitHub
    if group.is_external:
        repo_url = group.external_repo_url or group.project_link
        clone_dir = f"/tmp/rankbot_g{group.group}"
        path_note = f"\n- Sub-path to inspect: `{group.path}`" if group.path else ""
        return header + (
            f"This project is in an external GitHub repository.\n"
            f"- Repository URL: {repo_url}{path_note}\n\n"
            f"Inspect the project by running:\n"
            f"  [ -d {clone_dir} ] || git clone --depth 1 {repo_url} {clone_dir}\n"
            f"Then explore the files under `{clone_dir}/`.\n"
            f"Look for README, app entry points, agent definitions, and graph files."
        )

    match (group.branch, group.is_zip, group.is_commit):
        case (None, _, _):
            return header + (
                "This group has no submission link. "
                "Score 0/10 and explain that no code was available for review."
            )
        case (_, True, _):
            assert group.branch is not None
            assert group.path is not None
            return header + (
                f"This group submitted a .zip file.\n"
                f"- Repo: '{repo}'\n"
                f"- Branch: '{group.branch}'\n"
                f"- Zip path: '{group.path}'\n\n"
                f"You are running from the submissions repository workspace. "
                f"Use shell commands to inspect the zip without changing branches. "
                f"For example, run `git show origin/{group.branch}:{group.path}` "
                f"to read the zip bytes into a temporary file, then use `unzip -l` "
                f"to list files and `unzip -p` to read key files like README, "
                f"main app files, and agent/graph definitions.\n\n"
                f"Note: .zip submissions indicate poor code quality practices "
                f"(should have been committed properly to git)."
            )
        case (_, _, True):
            assert group.branch is not None
            path_hint = f"  Path hint: '{group.path}'" if group.path else ""
            return header + (
                f"This group's link points to a specific commit.\n"
                f"- Repo: '{repo}'\n"
                f"- Commit/Branch ref: '{group.branch}'\n"
                f"{path_hint}\n\n"
                f"You are running from the submissions repository workspace. "
                f"Use `git ls-tree -r --name-only {group.branch}` to list files, "
                f"then `git show {group.branch}:<filepath>` to read key files.\n"
                f"Look for README, app entry points, agent definitions, and graph files."
            )
        case _:
            assert group.branch is not None
            use_local = group.branch == "main"
            path_hint = group.path or ""
            path_clause = f" -- {path_hint}" if path_hint else ""

            if use_local:
                return header + (
                    f"This project is on the main branch.\n"
                    f"- Repo: '{repo}'\n"
                    f"- Directory: '{path_hint}'\n\n"
                    f"You are running from the submissions repository workspace. "
                    f"Use directory listings and file reads under `{path_hint}` "
                    f"to inspect the file structure and key files.\n"
                    f"Look for README, app entry points, agent definitions, and graph files."
                )
            return header + (
                f"This project is on a feature branch.\n"
                f"- Repo: '{repo}'\n"
                f"- Branch: '{group.branch}'\n"
                f"- Path: '{path_hint}'\n\n"
                f"You are running from the submissions repository workspace. "
                f"Use `git ls-tree -r --name-only origin/{group.branch}{path_clause}` "
                f"to see the file "
                f"structure, then `git show origin/{group.branch}:<filepath>` "
                f"to read key files.\n"
                f"Look for README, app entry points, agent definitions, and graph files."
            )


# ---------------------------------------------------------------------------
# Report generation
# ---------------------------------------------------------------------------


def generate_report(
    groups: list[GroupInfo],
    concept_scores: dict[int, ConceptScoreResult],
    difficulty_scores: dict[int, DifficultyScoreEntry],
    quality_scores: dict[int, CodeQualityResult],
    cohort: str = "C4",
) -> str:
    """Generate the final markdown evaluation report with rankings.

    Args:
        groups: All parsed group metadata.
        concept_scores: Concept scores keyed by group number.
        difficulty_scores: Difficulty scores keyed by group number.
        quality_scores: Code quality scores keyed by group number.
        cohort: Cohort label used in the report title (e.g. 'C4', 'C6').

    Returns:
        Complete markdown report string.
    """
    lines: list[str] = [
        f"# {cohort} Hackathon Evaluation Report",
        "",
        "## Summary",
        "",
        "| Rank | Group | Concept | Difficulty | Code Quality | Total |",
        "|------|-------|---------|------------|--------------|-------|",
    ]

    # Build rows with totals for sorting
    rows: list[tuple[int, int, int, int, int]] = []
    for g in groups:
        gn = g.group
        c = concept_scores.get(gn)
        d = difficulty_scores.get(gn)
        q = quality_scores.get(gn)
        cs = c.score if c else 0
        ds = d.score if d else 0
        qs = q.score if q else 0
        total = cs + ds + qs
        rows.append((gn, cs, ds, qs, total))

    rows.sort(key=lambda r: r[4], reverse=True)

    for rank, (gn, cs, ds, qs, total) in enumerate(rows, 1):
        lines.append(f"| {rank} | {gn} | {cs} | {ds} | {qs} | {total} |")

    lines.append("")
    lines.append("---")
    lines.append("")

    # Detailed per-group sections
    lines.append("## Detailed Evaluations")
    lines.append("")

    for gn, _cs_val, _ds_val, _qs_val, total in rows:
        lines.append(f"### Group {gn} (Total: {total}/30)")
        lines.append("")

        c = concept_scores.get(gn)
        if c:
            lines.append(f"**Concept Score: {c.score}/10**")
            lines.append(f"- Concepts found: {', '.join(c.concepts_found)}")
            lines.append(f"- Concepts missing: {', '.join(c.concepts_missing)}")
            lines.append(f"- Justification: {c.justification}")
        else:
            lines.append("**Concept Score: 0/10** — No submission")

        lines.append("")

        d = difficulty_scores.get(gn)
        if d:
            lines.append(f"**Difficulty Score: {d.score}/10**")
            lines.append(f"- Justification: {d.justification}")
        else:
            lines.append("**Difficulty Score: 0/10** — No submission")

        lines.append("")

        q = quality_scores.get(gn)
        if q:
            lines.append(f"**Code Quality Score: {q.score}/10**")
            lines.append(f"- Folder structure: {'✓' if q.has_proper_folders else '✗'}")
            lines.append(
                f"- README: {'✓' if q.has_readme else '✗'} ({q.readme_quality})"
            )
            lines.append(f"- Requirements: {'✓' if q.has_requirements_txt else '✗'}")
            lines.append(f"- Env handling: {'✓' if q.has_env_handling else '✗'}")
            lines.append(f"- Organization: {q.code_organization}")
            lines.append(f"- Justification: {q.justification}")
        else:
            lines.append("**Code Quality Score: 0/10** — No submission")

        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CSV writer — updates the scorecard CSV with computed scores
# ---------------------------------------------------------------------------


def write_scores_to_csv(
    csv_path: Path,
    concept_scores: dict[int, ConceptScoreResult],
    difficulty_scores: dict[int, DifficultyScoreEntry],
    quality_scores: dict[int, CodeQualityResult],
) -> None:
    """Read the scorecard CSV, fill in scores, compute totals and positions, write back.

    Updates columns: Concept Score (10), Difficulty Level (10),
    Code Quality (10), Total (30), Position.

    Args:
        csv_path: Path to the scorecard CSV to update in-place.
        concept_scores: Concept scores keyed by group number.
        difficulty_scores: Difficulty scores keyed by group number.
        quality_scores: Code quality scores keyed by group number.
    """
    # Read existing rows
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        assert fieldnames is not None, "CSV has no header row"
        rows = list(reader)

    # Fill in scores — only overwrite if we have a new score, preserve existing
    for row in rows:
        raw_group = row.get("Group", "").strip()
        if not raw_group or not raw_group.isdigit():
            continue
        gn = int(raw_group)

        c = concept_scores.get(gn)
        d = difficulty_scores.get(gn)
        q = quality_scores.get(gn)

        if c:
            row["Concept Score (10)"] = str(c.score)
        if d:
            row["Difficulty Level (10)"] = str(d.score)
        if q:
            row["Code Quality (10)"] = str(q.score)

        # Recompute total from whatever is in the row now
        cs = (
            int(row["Concept Score (10)"])
            if row.get("Concept Score (10)", "").strip().isdigit()
            else 0
        )
        ds = (
            int(row["Difficulty Level (10)"])
            if row.get("Difficulty Level (10)", "").strip().isdigit()
            else 0
        )
        qs = (
            int(row["Code Quality (10)"])
            if row.get("Code Quality (10)", "").strip().isdigit()
            else 0
        )
        total = cs + ds + qs
        row["Total (30)"] = str(total) if total > 0 else ""

    # Compute positions based on Total (descending)
    scored_rows = [
        (row, int(row["Total (30)"]))
        for row in rows
        if row.get("Total (30)", "").strip().isdigit()
    ]
    scored_rows.sort(key=lambda pair: pair[1], reverse=True)

    current_position = 0
    prev_total = None
    for i, (row, total) in enumerate(scored_rows):
        if total != prev_total:
            current_position = i + 1
        row["Position"] = str(current_position)
        prev_total = total

    # Write back
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    log.info("Updated CSV: %s", csv_path)


# ---------------------------------------------------------------------------
# xlsx writer — updates the scorecard xlsx with computed scores
# ---------------------------------------------------------------------------


def write_scores_to_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    concept_scores: dict[int, ConceptScoreResult],
    difficulty_scores: dict[int, DifficultyScoreEntry],
    quality_scores: dict[int, CodeQualityResult],
) -> None:
    """Write evaluation scores into the named sheet of an xlsx workbook.

    Preserves group row order exactly as-is. Fills Concept Score (10),
    Difficulty Level (10), Code Quality (10), Total (30), and Position columns.

    Args:
        xlsx_path: Path to the xlsx workbook to update in-place.
        sheet_name: Name of the sheet to update (e.g. 'C6').
        concept_scores: Concept scores keyed by group number.
        difficulty_scores: Difficulty scores keyed by group number.
        quality_scores: Code quality scores keyed by group number.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    # Locate header row and build column-name → column-index (1-based) map
    col_map: dict[str, int] = {}
    header_row_idx: int | None = None
    for row in ws.iter_rows():
        if row[0].value == "Group":
            header_row_idx = row[0].row
            for cell in row:
                if cell.value is not None:
                    col_map[str(cell.value)] = cell.column
            break

    assert (
        header_row_idx is not None
    ), f"No 'Group' header row found in sheet '{sheet_name}'"

    # First pass: write scores and collect totals for position calculation
    scored_totals: dict[int, int] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        gn = _parse_group_number(row[0].value)
        if gn is None:
            continue

        c = concept_scores.get(gn)
        d = difficulty_scores.get(gn)
        q = quality_scores.get(gn)
        row_idx = row[0].row

        if c and "Concept Score (10)" in col_map:
            ws.cell(row=row_idx, column=col_map["Concept Score (10)"]).value = c.score
        if d and "Difficulty Level (10)" in col_map:
            ws.cell(row=row_idx, column=col_map["Difficulty Level (10)"]).value = (
                d.score
            )
        if q and "Code Quality (10)" in col_map:
            ws.cell(row=row_idx, column=col_map["Code Quality (10)"]).value = q.score

        if "Comments" in col_map:
            parts: list[str] = []
            if c:
                parts.append(f"Concept: {c.justification}")
            if d:
                parts.append(f"Difficulty: {d.justification}")
            if q:
                parts.append(f"Quality: {q.justification}")
            if parts:
                ws.cell(row=row_idx, column=col_map["Comments"]).value = " | ".join(
                    parts
                )

        cs = c.score if c else 0
        ds = d.score if d else 0
        qs = q.score if q else 0
        total = cs + ds + qs
        if total > 0 and "Total (30)" in col_map:
            ws.cell(row=row_idx, column=col_map["Total (30)"]).value = total
            scored_totals[gn] = total

    # Compute positions by rank (highest total = position 1)
    sorted_groups = sorted(scored_totals.items(), key=lambda x: x[1], reverse=True)
    positions: dict[int, int] = {}
    prev_total = None
    current_pos = 0
    for i, (gn, total) in enumerate(sorted_groups):
        if total != prev_total:
            current_pos = i + 1
        positions[gn] = current_pos
        prev_total = total

    # Second pass: write positions without changing row order
    if "Position" in col_map:
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            gn = _parse_group_number(row[0].value)
            if gn is not None and gn in positions:
                ws.cell(row=row[0].row, column=col_map["Position"]).value = positions[
                    gn
                ]

    wb.save(xlsx_path)
    log.info("Updated xlsx: %s sheet=%s", xlsx_path, sheet_name)
